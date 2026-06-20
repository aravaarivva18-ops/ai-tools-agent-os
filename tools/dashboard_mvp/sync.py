from datetime import date, datetime, timedelta

import httpx
import openpyxl
from dashboard_mvp.models import (
    ChangeLog,
    DailyStat,
    Integration,
    KPIPlan,
    Project,
    SourceMapping,
)
from sqlalchemy.orm import Session


def get_mock_yandex_data(start_date_str: str, end_date_str: str):
    """Генерирует реалистичные mock-данные для Яндекс.Директ и Метрики."""
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()

    data = {}
    current_date = start_date

    # Генератор псевдослучайных чисел для имитации данных
    import random
    random.seed(current_date.toordinal())

    while current_date <= end_date:
        # В выходные трафик и конверсии чуть меньше
        is_weekend = current_date.weekday() in (5, 6)
        factor = 0.6 if is_weekend else 1.0

        impressions = int(random.randint(1000, 2500) * factor)
        clicks = int(impressions * random.uniform(0.05, 0.08)) # CTR 5-8%
        spent = round(clicks * random.uniform(15, 25), 2) # CPC 15-25 руб

        # Лиды с конверсией 3-5% от кликов
        leads = int(clicks * random.uniform(0.03, 0.06))
        # Округлим лиды до целого, минимум 0
        leads = max(0, leads)

        # Квалифицированные лиды (30-60% от лидов)
        qualified_leads = int(leads * random.uniform(0.3, 0.6))
        qualified_leads = max(0, qualified_leads)

        data[current_date] = {
            "impressions": impressions,
            "clicks": clicks,
            "spent": spent,
            "leads": leads,
            "qualified_leads": qualified_leads
        }
        current_date += timedelta(days=1)

    return data

async def sync_yandex_data(db: Session, project_id: int, start_date_str: str, end_date_str: str):
    """
    Синхронизирует данные из Яндекс.Директа и Яндекс.Метрики для проекта.
    Если токен отсутствует или равен 'mock_token', использует генератор mock-данных.
    """
    # 1. Получаем маппинг источников
    mapping = db.query(SourceMapping).filter(SourceMapping.project_id == project_id).first()
    if not mapping:
        print(f"Маппинг источников для проекта {project_id} не найден. Пропуск.")
        return False

    # 2. Получаем интеграцию с токеном
    integration = db.query(Integration).filter(
        Integration.project_id == project_id,
        Integration.type == "yandex"
    ).first()

    token = integration.access_token if integration else None

    # Режим имитации данных
    if not token or token == "mock_token":
        print(f"[SYNC MOCK] Синхронизация проекта {project_id} в режиме имитации.")
        mock_data = get_mock_yandex_data(start_date_str, end_date_str)
        project = db.query(Project).filter(Project.id == project_id).first()

        for dt, stats in mock_data.items():
            # Ищем или создаем запись статистики на этот день
            stat = db.query(DailyStat).filter(
                DailyStat.project_id == project_id,
                DailyStat.date == dt
            ).first()

            # Учет НДС
            spent = stats["spent"]
            if project and project.vat_type == "without_vat":
                spent = round(spent / (1.0 + project.vat_rate), 2)

            # Вычисляем производные показатели
            ctr = (stats["clicks"] / stats["impressions"] * 100) if stats["impressions"] > 0 else 0
            cpc = (spent / stats["clicks"]) if stats["clicks"] > 0 else 0
            cpl = (spent / stats["leads"]) if stats["leads"] > 0 else 0
            cpl_qualified = (spent / stats["qualified_leads"]) if stats["qualified_leads"] > 0 else 0

            if stat:
                stat.impressions = stats["impressions"]
                stat.clicks = stats["clicks"]
                stat.spent = spent
                stat.leads = stats["leads"]
                stat.qualified_leads = stats["qualified_leads"]
                stat.ctr = round(ctr, 2)
                stat.cpc = round(cpc, 2)
                stat.cpl = round(cpl, 2)
                stat.cpl_qualified = round(cpl_qualified, 2)
            else:
                stat = DailyStat(
                    project_id=project_id,
                    date=dt,
                    impressions=stats["impressions"],
                    clicks=stats["clicks"],
                    spent=spent,
                    leads=stats["leads"],
                    qualified_leads=stats["qualified_leads"],
                    ctr=round(ctr, 2),
                    cpc=round(cpc, 2),
                    cpl=round(cpl, 2),
                    cpl_qualified=round(cpl_qualified, 2)
                )
                db.add(stat)

        db.commit()
        return True

    # --- Реальный запрос к API Яндекса (при наличии токена) ---
    print(f"[SYNC] Запуск реальной синхронизации для проекта {project_id} по API.")
    project = db.query(Project).filter(Project.id == project_id).first()
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept-Language": "ru",
        "processingMode": "online"
    }
    if mapping.direct_login:
        headers["Client-Login"] = mapping.direct_login

    try:
        # 1. Запрос отчета в Директ API v5
        direct_url = "https://api.direct.yandex.com/json/v5/reports"
        report_query = {
            "params": {
                "SelectionCriteria": {
                    "DateFrom": start_date_str,
                    "DateTo": end_date_str
                },
                "FieldNames": ["Date", "Impressions", "Clicks", "Cost"],
                "ReportName": f"Daily_Report_{project_id}_{datetime.now().strftime('%M%S')}",
                "ReportType": "CUSTOM_REPORT",
                "DateRangeType": "CUSTOM_DATE",
                "Format": "TSV",
                "IncludeVAT": "YES",
                "IncludeDiscount": "NO"
            }
        }

        async with httpx.AsyncClient() as client:
            response = await client.post(direct_url, json=report_query, headers=headers, timeout=30.0)

        if response.status_code != 200:
            raise Exception(f"Ошибка API Директа: {response.status_code} - {response.text}")

        # Парсим TSV отчет Директа
        lines = response.text.strip().split('\n')
        # В отчете первая строка - название, вторая - заголовки, далее - данные
        direct_stats = {}
        for line in lines:
            parts = line.split('\t')
            if len(parts) >= 4 and parts[0] != "Date" and parts[0].count('-') == 2:
                dt_str = parts[0]
                dt = datetime.strptime(dt_str, "%Y-%m-%d").date()
                # Директ возвращает Cost в микро-валюте (умноженной на 1 000 000), делим
                spent = float(parts[3]) / 1000000.0 if parts[3].isdigit() else 0.0

                # Учет НДС
                if project and project.vat_type == "without_vat":
                    spent = spent / (1.0 + project.vat_rate)

                direct_stats[dt] = {
                    "impressions": int(parts[1]) if parts[1].isdigit() else 0,
                    "clicks": int(parts[2]) if parts[2].isdigit() else 0,
                    "spent": spent
                }

        # 2. Запрос конверсий из Метрики
        metrika_stats = {}
        if mapping.metrika_counter_id and (mapping.lead_goals_ids or mapping.qual_goals_ids):
            counter_id = mapping.metrika_counter_id
            goals = mapping.lead_goals_ids.split(',') if mapping.lead_goals_ids else []
            qual_goals = mapping.qual_goals_ids.split(',') if mapping.qual_goals_ids else []

            all_goals = goals + qual_goals
            goal_metrics = [f"ym:s:goal{goal_id}reaches" for goal_id in all_goals]
            metrics_param = ",".join(goal_metrics)

            metrika_url = (
                f"https://api-metrika.yandex.net/stat/v1/data?"
                f"ids={counter_id}"
                f"&metrics={metrics_param}"
                f"&dimensions=ym:s:date"
                f"&date1={start_date_str}"
                f"&date2={end_date_str}"
                f"&accuracy=full"
            )

            async with httpx.AsyncClient() as client:
                resp = await client.get(metrika_url, headers={"Authorization": f"Bearer {token}"}, timeout=30.0)

            if resp.status_code == 200:
                metrika_data = resp.json()
                # Парсим строки (rows) Метрики
                for row in metrika_data.get("data", []):
                    dt_str = row["dimensions"][0]["name"]
                    dt = datetime.strptime(dt_str, "%Y-%m-%d").date()
                    
                    # Разделяем метрики на обычные и квал-лиды
                    metrics_vals = [float(val) for val in row["metrics"]]
                    
                    leads_sum = sum(metrics_vals[i] for i in range(len(goals))) if goals else 0.0
                    qual_leads_sum = sum(metrics_vals[len(goals) + i] for i in range(len(qual_goals))) if qual_goals else 0.0
                    
                    metrika_stats[dt] = {
                        "leads": int(leads_sum),
                        "qualified_leads": int(qual_leads_sum)
                    }

        # 3. Объединяем и сохраняем данные в БД
        all_dates = set(direct_stats.keys()).union(set(metrika_stats.keys()))
        for dt in all_dates:
            d_stat = direct_stats.get(dt, {"impressions": 0, "clicks": 0, "spent": 0.0})
            m_stat = metrika_stats.get(dt, {"leads": 0, "qualified_leads": 0})

            stat = db.query(DailyStat).filter(
                DailyStat.project_id == project_id,
                DailyStat.date == dt
            ).first()

            ctr = (d_stat["clicks"] / d_stat["impressions"] * 100) if d_stat["impressions"] > 0 else 0
            cpc = (d_stat["spent"] / d_stat["clicks"]) if d_stat["clicks"] > 0 else 0
            cpl = (d_stat["spent"] / m_stat["leads"]) if m_stat["leads"] > 0 else 0
            cpl_qualified = (d_stat["spent"] / m_stat["qualified_leads"]) if m_stat["qualified_leads"] > 0 else 0

            if stat:
                stat.impressions = d_stat["impressions"]
                stat.clicks = d_stat["clicks"]
                stat.spent = d_stat["spent"]
                stat.leads = m_stat["leads"]
                stat.qualified_leads = m_stat["qualified_leads"]
                stat.ctr = round(ctr, 2)
                stat.cpc = round(cpc, 2)
                stat.cpl = round(cpl, 2)
                stat.cpl_qualified = round(cpl_qualified, 2)
            else:
                stat = DailyStat(
                    project_id=project_id,
                    date=dt,
                    impressions=d_stat["impressions"],
                    clicks=d_stat["clicks"],
                    spent=d_stat["spent"],
                    leads=m_stat["leads"],
                    qualified_leads=m_stat["qualified_leads"],
                    ctr=round(ctr, 2),
                    cpc=round(cpc, 2),
                    cpl=round(cpl, 2),
                    cpl_qualified=round(cpl_qualified, 2)
                )
                db.add(stat)

        db.commit()
        return True

    except Exception as e:
        db.rollback()
        print(f"Ошибка при синхронизации API для проекта {project_id}: {e}")
        return False

def sync_excel_data(db: Session, file_path: str):
    """Считывает планы KPI и логи изменений из заполненного Excel файла и импортирует в БД."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Не удалось открыть файл Excel: {e}")
        return False

    # --- 1. Импорт Плана KPI ---
    if "04_План_KPI" in wb.sheetnames:
        sheet = wb["04_План_KPI"]
        # Первая строка - заголовки
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]: # Если строка пустая
                continue
            month = str(row[0]) # Месяц (например, '2026-06')
            project_id_str = row[2]
            budget_plan = float(row[4]) if row[4] is not None else 0.0
            leads_plan = int(row[5]) if row[5] is not None else 0
            cpl_plan = float(row[6]) if row[6] is not None else 0.0
            qualified_leads_plan = int(row[7]) if (len(row) > 7 and row[7] is not None) else 0
            cpl_qualified_plan = float(row[8]) if (len(row) > 8 and row[8] is not None) else 0.0

            # Ищем проект по строковому ID (например PR-001 -> id 1)
            # В MVP просто выделим числовой ID из строки типа PR-001
            try:
                project_id = int(project_id_str.split('-')[-1]) if '-' in str(project_id_str) else int(project_id_str)
            except ValueError:
                continue

            # Проверяем, существует ли проект
            project_exists = db.query(Project.id).filter(Project.id == project_id).first() is not None
            if not project_exists:
                continue

            # Ищем существующий план
            plan = db.query(KPIPlan).filter(
                KPIPlan.project_id == project_id,
                KPIPlan.month == month
            ).first()

            if plan:
                plan.budget_plan = budget_plan
                plan.leads_plan = leads_plan
                plan.cpl_plan = cpl_plan
                plan.qualified_leads_plan = qualified_leads_plan
                plan.cpl_qualified_plan = cpl_qualified_plan
            else:
                plan = KPIPlan(
                    month=month,
                    project_id=project_id,
                    budget_plan=budget_plan,
                    leads_plan=leads_plan,
                    cpl_plan=cpl_plan,
                    qualified_leads_plan=qualified_leads_plan,
                    cpl_qualified_plan=cpl_qualified_plan
                )
                db.add(plan)
        db.commit()
        print("[SYNC EXCEL] Планы KPI успешно импортированы.")

    # --- 2. Импорт Лога изменений ---
    if "05_Лог_изменений" in wb.sheetnames:
        sheet = wb["05_Лог_изменений"]
        # Очищаем старые логи перед импортом, чтобы избежать дублирования
        # Для MVP просто будем добавлять только те, которых нет по дате и описанию
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            # Дата может быть datetime.date или строкой
            row_date = row[0]
            if isinstance(row_date, str):
                try:
                    dt = datetime.strptime(row_date, "%Y-%m-%d").date()
                except ValueError:
                    continue
            elif isinstance(row_date, (datetime, date)):
                dt = row_date if isinstance(row_date, date) else row_date.date()
            else:
                continue

            project_id_str = row[2]
            try:
                project_id = int(project_id_str.split('-')[-1]) if '-' in str(project_id_str) else int(project_id_str)
            except ValueError:
                continue

            changes = str(row[4])
            expected_effect = str(row[6]) if row[6] is not None else ""
            comment = str(row[7]) if row[7] is not None else ""

            # Проверяем проект
            if not db.query(Project.id).filter(Project.id == project_id).first() is not None:
                continue

            # Проверяем, есть ли лог
            log_exists = db.query(ChangeLog.id).filter(
                ChangeLog.project_id == project_id,
                ChangeLog.date == dt,
                ChangeLog.changes_description == changes
            ).first() is not None

            if not log_exists:
                log = ChangeLog(
                    date=dt,
                    project_id=project_id,
                    changes_description=changes,
                    comment=comment,
                    expected_effect=expected_effect
                )
                db.add(log)
        db.commit()
        print("[SYNC EXCEL] Логи изменений успешно импортированы.")

    return True
