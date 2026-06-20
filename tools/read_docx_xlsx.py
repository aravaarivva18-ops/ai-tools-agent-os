import sys

import docx
import openpyxl


def read_docx(file_path):
    try:
        doc = docx.Document(file_path)
    except Exception as e:
        return f"Ошибка при открытии DOCX: {e}"

    content = []
    # Читаем параграфы и таблицы по порядку их следования
    # Для этого перебираем элементы body
    for element in doc.element.body:
        if element.tag.endswith('p'):
            para = docx.text.paragraph.Paragraph(element, doc)
            if para.text.strip():
                # Попробуем определить заголовок
                if para.style.name.startswith('Heading'):
                    try:
                        level = int(para.style.name.split()[-1])
                        content.append("#" * level + " " + para.text)
                    except ValueError:
                        content.append("## " + para.text)
                else:
                    content.append(para.text)
        elif element.tag.endswith('tbl'):
            table = docx.table.Table(element, doc)
            content.append("\n[Таблица]")
            # Сделаем вывод в markdown-подобном виде
            for row in table.rows:
                # Избегаем дублирования ячеек при объединении (merge)
                cells = []
                for cell in row.cells:
                    text = cell.text.strip().replace('\n', ' ')
                    cells.append(text)
                content.append("| " + " | ".join(cells) + " |")
            content.append("")

    return "\n".join(content)

def read_xlsx(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return f"Ошибка при открытии XLSX: {e}"

    content = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        content.append(f"\n================ Лист: {sheet_name} ================")

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            content.append("(Лист пустой)")
            continue

        # Удаляем полностью пустые строки в конце
        while rows and all(val is None for val in rows[-1]):
            rows.pop()

        for idx, row in enumerate(rows):
            # Проверяем, пустая ли строка
            if all(val is None for val in row):
                continue
            # Форматируем ячейки (округляем флоаты, форматируем None)
            formatted_cells = []
            for cell in row:
                if cell is None:
                    formatted_cells.append("")
                elif isinstance(cell, float):
                    # Если число целое по сути, уберем .0
                    if cell.is_integer():
                        formatted_cells.append(str(int(cell)))
                    else:
                        formatted_cells.append(f"{cell:.4f}".rstrip('0').rstrip('.'))
                else:
                    formatted_cells.append(str(cell))
            content.append(f"Строка {idx+1}: " + " | ".join(formatted_cells))

    return "\n".join(content)

def main():
    if len(sys.argv) < 3:
        print("Использование: python read_docx_xlsx.py <тип: docx|xlsx> <путь_к_файлу>")
        sys.exit(1)

    file_type = sys.argv[1].lower()
    file_path = sys.argv[2]

    if file_type == 'docx':
        print(read_docx(file_path))
    elif file_type == 'xlsx':
        print(read_xlsx(file_path))
    else:
        print("Неизвестный тип. Используйте 'docx' или 'xlsx'.")
        sys.exit(1)

if __name__ == '__main__':
    main()
