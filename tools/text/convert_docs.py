"""Utility script to convert various document formats to Markdown using Microsoft MarkItDown."""

import argparse
import sys
from pathlib import Path


def convert_document(input_path: Path, output_path: Path | None = None) -> None:
    """Convert a document file to clean Markdown using MarkItDown or fallbacks.

    Args:
        input_path: Path to the input file (PDF, Docx, etc.).
        output_path: Optional path to save the converted Markdown file.
            If not provided, the output is written to standard output.
    """
    if not input_path.exists():
        print(f"Error: The input file '{input_path}' does not exist.", file=sys.stderr)
        sys.exit(1)

    markdown_text = ""
    try:
        # Try using Microsoft MarkItDown first
        try:
            from markitdown import MarkItDown

            markitdown = MarkItDown()
            result = markitdown.convert(str(input_path))
            markdown_text = result.text_content
        except ImportError:
            # Fallback for Python 3.14 (where onnxruntime/markitdown is unsupported)
            suffix = input_path.suffix.lower()
            if suffix == ".pdf":
                import pypdf

                reader = pypdf.PdfReader(input_path)
                pages_text = []
                for i, page in enumerate(reader.pages):
                    text = page.extract_text() or ""
                    pages_text.append(f"## Page {i + 1}\n\n{text}")
                markdown_text = "\n\n".join(pages_text)
            elif suffix == ".docx":
                import docx

                doc = docx.Document(str(input_path))
                paras = [p.text for p in doc.paragraphs]
                markdown_text = "\n\n".join(paras)
            elif suffix in (".xlsx", ".xls"):
                import openpyxl

                wb = openpyxl.load_workbook(input_path, data_only=True)
                sheets = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        if any(row):
                            rows.append(
                                " | ".join(str(v) if v is not None else "" for v in row)
                            )
                    sheets.append(f"## Sheet: {sheet_name}\n\n" + "\n".join(rows))
                markdown_text = "\n\n".join(sheets)
            else:
                raise ValueError(
                    f"Unsupported format '{suffix}' for fallback parsing."
                ) from None

        if output_path:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(markdown_text, encoding="utf-8")
            print(f"Successfully converted '{input_path}' to '{output_path}'")
        else:
            print(markdown_text)
    except Exception as e:
        print(f"Error during document conversion: {e}", file=sys.stderr)
        sys.exit(1)


def main() -> None:
    """Parse command line arguments and execute the document conversion."""
    parser = argparse.ArgumentParser(
        description="Convert documents (PDF, DOCX, XLSX, etc.) to clean Markdown."
    )
    parser.add_argument(
        "input",
        type=str,
        help="Path to the input document file.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=str,
        default=None,
        help="Optional path to write the converted Markdown file.",
    )

    args = parser.parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else None

    convert_document(input_path, output_path)


if __name__ == "__main__":
    main()
